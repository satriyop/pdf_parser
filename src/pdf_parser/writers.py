import csv
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class CsvWriter:
    def __init__(self, output_path):
        self.output_path = output_path

    def write(self, site_data_list, sheet_name=""):
        if not site_data_list:
            return
        headers = site_data_list[0].csv_headers()
        file_exists = os.path.isfile(self.output_path)

        with open(self.output_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(headers)
            for site in site_data_list:
                writer.writerow(site.to_csv_row())


class XlsxWriter:
    def __init__(self, output_path):
        self.output_path = output_path

    def write(self, site_data_list, sheet_name="Survey Data"):
        if not site_data_list:
            return

        if os.path.isfile(self.output_path):
            wb = load_workbook(self.output_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(site_data_list[0].csv_headers())
            self._style_header(ws, len(site_data_list[0].csv_headers()))

        for site in site_data_list:
            ws.append(site.to_xlsx_row())

        self._auto_width(ws)
        wb.save(self.output_path)

    def _style_header(self, ws, col_count):
        bold = Font(bold=True)
        for col in range(1, col_count + 1):
            ws.cell(row=1, column=col).font = bold

    def _auto_width(self, ws):
        for col_cells in ws.columns:
            try:
                col_letter = get_column_letter(col_cells[0].column)
                max_len = max(len(str(c.value or "")) for c in col_cells)
                ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
            except Exception:
                pass
