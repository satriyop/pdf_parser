import csv
import os
import tempfile

import pytest
from openpyxl import load_workbook

from pdf_parser.models import SiteData
from pdf_parser.writers import CsvWriter, XlsxWriter


class TestCsvWriter:
    def test_write_new_file(self, tmp_path, sample_sites):
        out = tmp_path / "test.csv"
        writer = CsvWriter(str(out))
        writer.write(sample_sites)

        rows = list(csv.reader(open(out)))
        assert len(rows) == 3
        assert rows[0] == SiteData.csv_headers()
        assert rows[1][2] == "SITE_ALPHA"
        assert rows[2][2] == "SITE_BETA"

    def test_append_to_existing(self, tmp_path, sample_sites):
        out = tmp_path / "test.csv"
        writer = CsvWriter(str(out))
        writer.write(sample_sites[:1])
        writer.write(sample_sites[1:])

        rows = list(csv.reader(open(out)))
        assert len(rows) == 3
        assert rows[0] == SiteData.csv_headers()

    def test_write_empty_list(self, tmp_path):
        out = tmp_path / "test.csv"
        writer = CsvWriter(str(out))
        writer.write([])
        assert not out.exists() or os.path.getsize(out) == 0


class TestXlsxWriter:
    def test_write_new_file(self, tmp_path, sample_sites):
        out = tmp_path / "test.xlsx"
        writer = XlsxWriter(str(out))
        writer.write(sample_sites, sheet_name="EAST JAVA")

        wb = load_workbook(out)
        assert "EAST JAVA" in wb.sheetnames
        ws = wb["EAST JAVA"]
        assert ws.cell(1, 1).value == "NO"
        assert ws.cell(2, 2).value == "EAST JAVA"
        assert ws.cell(3, 2).value == "EAST JAVA"

    def test_append_to_existing_sheet(self, tmp_path, sample_sites):
        out = tmp_path / "test.xlsx"
        writer = XlsxWriter(str(out))
        writer.write(sample_sites[:1], sheet_name="DATA")
        writer.write(sample_sites[1:], sheet_name="DATA")

        wb = load_workbook(out)
        ws = wb["DATA"]
        assert ws.cell(1, 1).value == "NO"
        assert ws.cell(2, 2).value == "EAST JAVA"
        assert ws.cell(3, 2).value == "EAST JAVA"

    def test_multiple_sheets(self, tmp_path, sample_sites):
        out = tmp_path / "test.xlsx"
        writer = XlsxWriter(str(out))
        writer.write([sample_sites[0]], sheet_name="AREA_A")
        writer.write([sample_sites[1]], sheet_name="AREA_B")

        wb = load_workbook(out)
        assert "AREA_A" in wb.sheetnames
        assert "AREA_B" in wb.sheetnames

    def test_write_empty_list(self, tmp_path):
        out = tmp_path / "test.xlsx"
        writer = XlsxWriter(str(out))
        writer.write([])
        assert not out.exists()

    def test_headers_bold(self, tmp_path, sample_sites):
        out = tmp_path / "test.xlsx"
        writer = XlsxWriter(str(out))
        writer.write(sample_sites[:1], sheet_name="S1")
        wb = load_workbook(out)
        ws = wb["S1"]
        assert ws.cell(1, 1).font.bold is True
